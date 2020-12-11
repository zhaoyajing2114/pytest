import requests
import pytest
import json
import openpyxl
import os
'''
to_dos: 将解析完的请求结构存放到excel中，方便拓展和增加用例
'''
#解析har文件
def decode_har(cookie):
    api_file = openpyxl.Workbook()
    with open(r'D:\workspace\add1.har', 'r', encoding='utf-8') as har_file:
        har_content = har_file.read()
        if har_content.startswith('\ufeff'):#导出的har文件开头都包含\ufeff  UTF-8 BOM (decode using utf-8-sig)
            har_content = har_content.encode('utf-8')[3:].decode('utf-8')
        har_json = json.loads(har_content)
        object_list = har_json['log']['entries']
        api_num = 1
        for object in object_list:
            request_json = object['request']
            response_json = object['response']
            header_list = request_json['headers']
            header_dict = {}
            for header_content in header_list:#获取请求头
                name = header_content['name']
                value = header_content['value']
                header_dict.update({name:value})
            # header_dict.update({'cookies':str(cookie)})
            method = request_json['method']#获取请求方法
            url_str = request_json['url']#请求路径
            response_status = response_json['status']
            resopnse_txt = response_json['content']['text']
            resopnse_type = response_json['content']['mimeType']
            if method == 'POST':
                mime_type = request_json['postData']['mimeType']
                post_data_str = request_json['postData']['text']
                if mime_type == 'application/json':
                    post_data = json.loads(post_data_str)#post请求的请求内容
                    if 'summary' not in api_file.sheetnames:
                        sht = api_file.create_sheet('summary')
                        sht.cell(1, 1).value = '页面名称'
                        sht.cell(1, 2).value = 'url'
                        sht.cell(1, 3).value = 'execute'
                    else:
                        sht = api_file['summary']
                    sht.cell(1+api_num,1).value = 'api'+str(api_num)
                    sht.cell(1+api_num,2).value = url_str
                    sht.cell(1 + api_num, 3).value = 'Y'
                    sht = api_file.create_sheet('api'+str(len(api_file.sheetnames)))
                    sht.cell(1,1).value = url_str
                    sht.cell(2,1).value = method
                    sht.cell(3,1).value = str(header_dict)
                    api_num += 1
                    i = 4
                    for item in post_data:
                        sht.cell(i, 2).value = item
                        sht.cell(i, 3).value = post_data[item]
                        i += 1
                    sht.cell(i,2).value = 'status'
                    sht.cell(i,3).value = response_status
                    sht.cell(i + 1, 2).value = 'type'
                    sht.cell(i + 1, 3).value = resopnse_type
                    sht.cell(i + 2, 2).value = 'text'
                    sht.cell(i + 2, 3).value = resopnse_txt
                else:
                    sht = api_file.get_sheet_by_name['summary']
                    result = requests.post(url_str,headers=header_dict,json=post_data)
                    print(result.text)
            elif method == 'GET':
                try:
                    result = requests.get(url_str, headers=header_dict)
                    print(result.text)
                    if 'summary' not in api_file.sheetnames:
                        sht = api_file.create_sheet('summary')
                        sht.cell(1, 1).value = '页面名称'
                        sht.cell(1, 2).value = 'url'
                        sht.cell(1, 3).value = 'execute'
                    else:
                        sht = api_file['summary']
                    sht.cell(1 + api_num, 1).value = 'api' + str(api_num)
                    sht.cell(1 + api_num, 2).value = url_str
                    sht.cell(1 + api_num, 3).value = 'Y'
                    sht = api_file.create_sheet('api' + str(len(api_file.sheetnames)))#
                    sht.cell(1, 1).value = url_str
                    sht.cell(2, 1).value = method
                    sht.cell(3, 1).value = str(header_dict)
                    sht.cell(4, 2).value = 'status'
                    sht.cell(4, 3).value = response_status
                    sht.cell(5, 2).value = 'type'
                    sht.cell(5, 3).value = resopnse_type
                    sht.cell(6, 2).value = 'text'
                    sht.cell(6, 3).value = resopnse_txt
                    api_num += 1
                except BaseException as e:
                    print(str(e))
        api_file.save(r'D:\workspace\pytest1\api_excel\test.xlsx')


if __name__ == '__main__':
    decode_har('1')