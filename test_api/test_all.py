import pytest
import requests
import os
import openpyxl
from common.request_method import requests_get,requests_post

def setup_module(module):
    print('setup module')

def teardown_module(module):
    print('teardown module')

def test_all(get_conf):
    if os.path.exists(get_conf['case_path']):
        wkbk = openpyxl.load_workbook(get_conf['case_path'])
        summary_sht = wkbk['summary']
        row_num = summary_sht.max_row
        for i in range(2,row_num):
            print(summary_sht.cell(i,3).value)
            execute_str = summary_sht.cell(i,3).value
            api_name = summary_sht.cell(i, 1).value
            if execute_str == 'Y':
                cur_sht = wkbk[api_name]
                url = cur_sht.cell(1,1).value
                method = cur_sht.cell(2,1).value
                header = cur_sht.cell(3,1).value
                print(url)
                print(method)
                print(header)
                if method == 'POST':
                    for i in range(5,100):
                        print(i)
            else:
                print(api_name+'不需要进行测试')

    else:
        print('用例所在路径不存在')
